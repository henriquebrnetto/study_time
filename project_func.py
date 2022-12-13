# -*- coding: utf-8 -*-
"""
@author: Henrique Bucci

"""
import math
import pandas as pd
import numpy as np
import mysql.connector as mysql
import csv
import os, re
from datetime import datetime, timedelta, date

def find(name, path):
    for root, dirs, files in os.walk(path):
        if name in files:
            return os.path.join(root, name)
        

#Time-Step between Measurements
def diff(data, interval):
    time = []
    for t in range(len(data)):
        if t%interval == 0 and t != 0:
            time.append(t)
    return [(data[t] - data[t-interval]).total_seconds() for t in time]


def passed_time(data, interval):
    t = [0] 
    time = 0 #The passed time of the first measurement should be used as the "beginning", therefore it should be equal to 0.
    #if type(data[0]) == 'pandas._libs.tslibs.timestamps.Timestamp' or type(data[0]) == 'datetime' or type(data[0]) == 'timedelta':
    for i in range(interval, len(data)):
        time = time + (data[i] - data[i-interval]).total_seconds()
        t.append(time)
    return t 
   

#Splitting X-axis Train/Test Data (from same dataframe)
def xtrain_test(df, feats = None, size=0.75):
    size = int(size*len(df))
    if feats == None:
        train = df.loc[:size-1]
        test = df.loc[size:]
    else:
        train = df.loc[:size-1, feats]
        test = df.loc[size:, feats]
    return train, test


def ytrain_test(df, col = -1, size=0.75):
    size = int(size*len(df))

    #Train Target
    train = np.array([float(x) for x in df.loc[:size-1, col]]).reshape(len(test),1)

    #Test Target
    test = np.array([float(x) for x in df.loc[size:,col]]).reshape(len(test),1)
    return train, test

def sql_to_df(cursor, query, *col):
    cursor.execute(query)
    if len(col) == 1:
        return pd.DataFrame([x[0] for x in cursor], columns=col)
    elif len(col) == 0:
        return pd.DataFrame([x for x in cursor])
    else:
        return pd.DataFrame([x for x in cursor], columns=col)
    

def sql_to_csv(cursor, query, **kwargs):
    cursor.execute(query)
    col = kwargs.get('col', []) #def of col kwarg
    filename = kwargs.get('filename', None) # def of filename kwarg
    if filename != None:    #called function has filename argument
        if os.path.exists(f'{filename}.csv') == True:  #if the file exists
            with  open(f'{filename}.csv', 'a', newline='') as file:
                csv_out = csv.writer(file)
                [csv_out.writerow(line) for line in cursor]
            return
        else:   #if the file does not exist
            try:
                with open(f'{filename}.csv', 'w', newline='') as file:
                    csv_out = csv.writer(file)
                    if len(col) == 0:   #called function with no col argument
                        csv_out.writerow('0')
                        [csv_out.writerow(line) for line in cursor]
                    elif len(col) == 1:
                        csv_out.writerow(col)
                        [csv_out.writerow(line) for line in cursor]
                    else:
                        csv_out.writerow(col)
                        [csv_out.writerow(line) for line in cursor]
                    return
            except ValueError:
                raise TypeError("Column argument must be an iterable object")
    else:   #called function does not have filename argument
        try:
            if len(col) == 1:
                df = pd.DataFrame([x[0] for x in cursor], columns=[col])
            elif len(col) == 0:     #called function with no col argument
                df = pd.DataFrame([x for x in cursor])
            else:
                df = pd.DataFrame([x for x in cursor], columns=[col])
            return df
        except ValueError:
            raise TypeError("Column argument must be an iterable object")


def file_reader(*args, **kwargs):
    filetype = kwargs.get('filetype', None)
    try:
        if filetype == None:
            if args[0].__class__ == list:
                return [pd.read_csv(x) for x in args[0]]
            else:
                return [pd.read_csv(x) for x in args]
        elif filetype == 'xlsx' | filetype == 'excel':
            if args[0].__class__ == list:
                return [pd.read_excel(x) for x in args[0]]
            else:
                return [pd.read_excel(x) for x in args]
    except UnicodeDecodeError:
        if filetype == None:
            if args[0].__class__ == list:
                return [pd.read_csv(x, encoding='ISO-8859-1') for x in args[0]]
            else:
                return [pd.read_csv(x, encoding='ISO-8859-1') for x in args]
        elif filetype == 'xlsx' | filetype == 'excel':
            if args[0].__class__ == list:
                return [pd.read_excel(x, encoding='ISO-8859-1') for x in args[0]]
            else:
                return [pd.read_excel(x, encoding='ISO-8859-1') for x in args]
    

def csv_append(cursor, col, table, **kwargs):
    filename = kwargs.get('filename', col)
    where = kwargs.get('where', '')
    query = kwargs.get('query', None)
    if filename.endswith('.csv') == True:
        df = pd.read_csv(filename)
    else:
        df = pd.read_csv(f'{filename}.csv')
    size = len(df.iloc[:,0])
    if query == None:
        if where == '':
            quer = f'SELECT {col} FROM {table};'
        else:
            quer = f'SELECT {col} FROM {table} WHERE {where};'
        cursor.execute(quer)
    else:
        cursor.execute(query)
    sql = [x for x in cursor]
    sql = sql[size:]
    if os.path.exists(f'{filename}.csv') == True:  #if the file exists
        with  open(f'{filename}.csv', 'a', newline='') as file:
            csv_out = csv.writer(file)
            [csv_out.writerow(line) for line in sql]
        return
    else:
        with open(f'{filename}.csv', 'w', newline='') as file:
            csv_out = csv.writer(file)
            df = [csv_out.writerow(line) for line in sql]
        return
    

"""
def insert_sql(cursor, table, data, cols = None):
    n = data.shape[1]
    size = ','.join(['%s' for x in range(n)])
    if cols == None and type(data) != 'pandas.core.frame.DataFrame':
        raise Exception('cols argument must be different than None or data argument must have column names')
    if cols == None:
        cols = (data.columns)
    if len(cols) > 1:
        cols = ','.join(cols)
    for index, row in data.iterrows():
        val = [row[x] for x in range(n)]
        query = ('INSERT INTO %s (%s) VALUES (%s)' % (table, cols, size))
        cursor.execute(query, val)
        db.commit()
    return    
"""
    
def create_book(df_names=None, *, first_sheet = None, filename='newfile', data_only = False):
    from openpyxl import Workbook, load_workbook
    import os
    if os.path.isfile(f'{filename}.xlsx') == False:
        wb = Workbook()
        if first_sheet == None:
            sheet = wb.active
            sheet.title = df_names[0]
            del df_names[0]
        else:
            sheet = wb.active
            sheet.title = first_sheet
        if df_names != None:
            for df in df_names:
                sheet = wb.create_sheet(df)
        wb.save(f'{filename}.xlsx')
        ws = wb.get_sheet_names()
    #If file already exists, it is possible to provide only 'filename' variable
    else:
        wb = load_workbook(f'{filename}.xlsx', data_only=data_only)
        ws = wb.get_sheet_names()
    return wb, ws

"""def add_sheet(new_sheets, *, filename='newfile', pos = -1):
    from openpyxl import load_workbook
    if type(new_sheets) != list:
        raise f'{TypeError} : Variable must have type {list}'
    wb = load_workbook(f'{filename}.xlsx')
    ws = wb.get_sheet_names()

    #Get data from existing sheets
    for sheet in wb:
        size = sheet.dimensions

    for sheet in new_sheets:
        ws.insert(pos, sheet)
    sheet = wb.active
    sheet.title = new_sheets[0]
    del new_sheets[0]
    for sheet in new_sheets:
        sheet = wb.create_sheet(sheet)"""
    
    
    



    
    
    
    
    
    
    
    
    
    
    
    